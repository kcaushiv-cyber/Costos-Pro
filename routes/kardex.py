from flask import Blueprint, render_template, session, redirect, url_for, request, jsonify, flash, send_file, current_app
from models.database import query_db, execute_db
from routes.dashboard import login_required, company_required
from datetime import date, datetime
import io

kardex_bp = Blueprint('kardex', __name__)


def _check_safety_stock(item_id, company_id, new_stock):
    """Verifica si el stock cae bajo el mínimo y envía alerta por correo."""
    # JOIN con units_of_measure para obtener unit_code
    item = query_db(
        """SELECT ii.*, COALESCE(u.code, 'uds') as unit_code
           FROM inventory_items ii
           LEFT JOIN units_of_measure u ON ii.unit_id = u.id
           WHERE ii.id=? AND ii.company_id=?""",
        (item_id, company_id), one=True
    )
    if not item or not item['safety_stock'] or float(item['safety_stock']) <= 0:
        return
    if new_stock < float(item['safety_stock']):
        try:
            from utils.email_utils import send_low_stock_email
            company = query_db(
                """SELECT c.business_name, u.email, u.full_name
                   FROM companies c JOIN users u ON c.user_id = u.id
                   WHERE c.id=?""",
                (company_id,), one=True
            )
            if not company or not company['email']:
                current_app.logger.warning(f'Sin email de empresa para alerta stock item {item_id}')
                return
            mail = current_app.extensions.get('mail')
            if not mail or not current_app.config.get('MAIL_USERNAME'):
                current_app.logger.warning('Mail no configurado — no se envía alerta stock')
                return
            send_low_stock_email(
                mail,
                to_email=company['email'],
                full_name=company['full_name'],
                company_name=company['business_name'],
                item_code=item['code'],
                item_name=item['name'],
                current_stock=new_stock,
                safety_stock=float(item['safety_stock']),
                unit=item['unit_code'] or 'uds'
            )
            current_app.logger.info(f'Alerta stock enviada: {item["name"]} → {company["email"]}')
        except Exception as e:
            current_app.logger.error(f'Error enviando alerta stock: {e}')


@kardex_bp.route('/kardex')
@login_required
@company_required
def index():
    session['active_module'] = 'kardex'
    company_id = session['company_id']
    search = request.args.get('search', '')
    cat_filter = request.args.get('category', '')

    q = """SELECT ii.*, u.code as unit_code, u.name as unit_name
           FROM inventory_items ii
           LEFT JOIN units_of_measure u ON ii.unit_id = u.id
           WHERE ii.company_id = ? AND ii.is_active = 1"""
    params = [company_id]
    if search:
        q += " AND (ii.name LIKE ? OR ii.code LIKE ?)"
        params += [f'%{search}%', f'%{search}%']
    if cat_filter:
        q += " AND ii.category = ?"
        params.append(cat_filter)
    q += " ORDER BY ii.name"

    items = query_db(q, params)
    total_value = sum((i['current_stock'] or 0) * (i['average_cost'] or 0) for i in items)
    units = query_db("SELECT * FROM units_of_measure WHERE company_id=? ORDER BY name", (company_id,))
    return render_template('kardex/index.html',
                           items=items, search=search,
                           cat_filter=cat_filter,
                           total_value=total_value,
                           units=units)


@kardex_bp.route('/kardex/new', methods=['GET', 'POST'])
@login_required
@company_required
def new_item():
    company_id = session['company_id']

    if request.method == 'POST':
        d = request.form
        name          = d.get('name', '').strip()
        code          = d.get('code', '').strip()
        category      = d.get('category', 'insumo')
        # Unidad: texto libre → buscar o crear en units_of_measure
        unit_id   = None
        unit_name = (d.get('unit_name') or '').strip()
        if unit_name:
            # Extraer código si viene como "Kilogramo (kg)"
            import re as _re
            m = _re.search(r'\(([^)]+)\)', unit_name)
            unit_code_guess = m.group(1).upper() if m else unit_name[:6].upper()
            clean_name = _re.sub(r'\s*\([^)]*\)', '', unit_name).strip() or unit_name

            existing_unit = query_db(
                "SELECT id FROM units_of_measure WHERE company_id=? AND (name=? OR code=?)",
                (company_id, clean_name, unit_code_guess), one=True
            )
            if existing_unit:
                unit_id = existing_unit['id']
            else:
                unit_id = execute_db(
                    "INSERT INTO units_of_measure (company_id, code, name, category) VALUES (?,?,?,?)",
                    (company_id, unit_code_guess, clean_name, 'otro')
                )
        initial_stock = float(d.get('initial_stock') or 0)
        initial_cost  = float(d.get('initial_cost') or 0)
        safety_stock  = float(d.get('safety_stock') or 0)
        valuation     = d.get('valuation_method', 'promedio')
        initial_date  = d.get('initial_date') or str(date.today())
        reference     = d.get('initial_reference', 'Saldo inicial').strip()

        if not name or not code:
            flash('Nombre y código son obligatorios.', 'error')
            units = query_db("SELECT * FROM units_of_measure WHERE company_id=? ORDER BY name", (company_id,))
            return render_template('kardex/new_item.html', units=units)

        existing = query_db(
            "SELECT id FROM inventory_items WHERE company_id=? AND code=?",
            (company_id, code), one=True
        )
        if existing:
            flash('Ya existe un ítem con ese código.', 'error')
            units = query_db("SELECT * FROM units_of_measure WHERE company_id=? ORDER BY name", (company_id,))
            return render_template('kardex/new_item.html', units=units)

        item_id = execute_db(
            """INSERT INTO inventory_items
               (company_id, code, name, category, unit_id,
                current_stock, average_cost, safety_stock, valuation_method, is_active)
               VALUES (?,?,?,?,?,?,?,?,?,1)""",
            (company_id, code, name, category, unit_id,
             initial_stock, initial_cost, safety_stock, valuation)
        )

        if initial_stock > 0:
            execute_db(
                """INSERT INTO kardex_movements
                   (inventory_item_id, movement_date, movement_type,
                    quantity, unit_cost, total_cost, stock_after, average_cost_after, reference)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (item_id, initial_date, 'entrada',
                 initial_stock, initial_cost,
                 initial_stock * initial_cost,
                 initial_stock, initial_cost, reference)
            )

        flash('Ítem creado exitosamente.', 'success')
        return redirect(url_for('kardex.item_detail', item_id=item_id))

    units = query_db("SELECT * FROM units_of_measure WHERE company_id=? ORDER BY name", (company_id,))
    return render_template('kardex/new_item.html', units=units)


@kardex_bp.route('/kardex/item/<int:item_id>')
@login_required
@company_required
def item_detail(item_id):
    company_id = session['company_id']

    item = query_db(
        """SELECT ii.*, u.code as unit_code, u.name as unit_name
           FROM inventory_items ii
           LEFT JOIN units_of_measure u ON ii.unit_id = u.id
           WHERE ii.id=? AND ii.company_id=?""",
        (item_id, company_id), one=True
    )
    if not item:
        flash('Ítem no encontrado.', 'error')
        return redirect(url_for('kardex.index'))

    movements = query_db(
        "SELECT * FROM kardex_movements WHERE inventory_item_id=? ORDER BY movement_date DESC, id DESC",
        (item_id,)
    )
    units = query_db("SELECT * FROM units_of_measure WHERE company_id=? ORDER BY name", (company_id,))
    return render_template('kardex/item_detail.html', item=item, movements=movements, units=units)


@kardex_bp.route('/kardex/item/<int:item_id>/edit', methods=['POST'])
@login_required
@company_required
def edit_item(item_id):
    """Editar datos del ítem (nombre, categoría, stock mínimo, etc.)"""
    company_id = session['company_id']
    d = request.form

    # Unidad: texto libre → buscar o crear en units_of_measure
    unit_id   = None
    unit_name = (d.get('unit_name') or '').strip()
    if unit_name:
        import re as _re
        m = _re.search(r'\(([^)]+)\)', unit_name)
        unit_code_guess = m.group(1).upper() if m else unit_name[:6].upper()
        clean_name = _re.sub(r'\s*\([^)]*\)', '', unit_name).strip() or unit_name

        existing_unit = query_db(
            "SELECT id FROM units_of_measure WHERE company_id=? AND (name=? OR code=?)",
            (company_id, clean_name, unit_code_guess), one=True
        )
        if existing_unit:
            unit_id = existing_unit['id']
        else:
            unit_id = execute_db(
                "INSERT INTO units_of_measure (company_id, code, name, category) VALUES (?,?,?,?)",
                (company_id, unit_code_guess, clean_name, 'otro')
            )

    execute_db(
        """UPDATE inventory_items SET
           code=?, name=?, category=?, unit_id=?,
           safety_stock=?, valuation_method=?,
           updated_at=datetime('now')
           WHERE id=? AND company_id=?""",
        (d.get('code'), d.get('name'), d.get('category', 'insumo'),
         unit_id,
         float(d.get('safety_stock') or 0),
         d.get('valuation_method', 'promedio'),
         item_id, company_id)
    )
    flash('Ítem actualizado.', 'success')
    return redirect(url_for('kardex.item_detail', item_id=item_id))


@kardex_bp.route('/kardex/item/<int:item_id>/movement', methods=['POST'])
@login_required
@company_required
def add_movement(item_id):
    company_id = session['company_id']

    item = query_db(
        "SELECT * FROM inventory_items WHERE id=? AND company_id=?",
        (item_id, company_id), one=True
    )
    if not item:
        return jsonify({'error': 'No autorizado'}), 403

    data          = request.get_json()
    movement_type = data.get('movement_type')
    quantity      = float(data.get('quantity', 0) or 0)
    unit_cost     = float(data.get('unit_cost', 0) or 0)
    movement_date = data.get('movement_date', str(date.today()))
    reference     = data.get('reference', '')
    doc_type      = data.get('document_type', '')
    doc_number    = data.get('document_number', '')
    notes         = data.get('notes', '')

    if quantity <= 0:
        return jsonify({'error': 'La cantidad debe ser positiva.'}), 400

    current_stock = item['current_stock'] or 0
    current_avg   = item['average_cost'] or 0

    if movement_type == 'entrada':
        if current_stock + quantity > 0:
            new_avg = (current_stock * current_avg + quantity * unit_cost) / (current_stock + quantity)
        else:
            new_avg = unit_cost
        new_stock  = current_stock + quantity
        total_cost = quantity * unit_cost

    elif movement_type == 'salida':
        if quantity > current_stock:
            return jsonify({'error': f'Stock insuficiente. Disponible: {current_stock:.3f}'}), 400
        new_avg    = current_avg
        new_stock  = current_stock - quantity
        total_cost = quantity * current_avg
        unit_cost  = current_avg

    elif movement_type == 'devolucion':
        new_avg    = current_avg
        new_stock  = current_stock + quantity
        total_cost = quantity * current_avg

    elif movement_type == 'ajuste':
        new_stock  = quantity
        new_avg    = unit_cost if unit_cost > 0 else current_avg
        total_cost = new_stock * new_avg
        quantity   = new_stock - current_stock

    else:
        return jsonify({'error': 'Tipo de movimiento inválido.'}), 400

    execute_db(
        """INSERT INTO kardex_movements
           (inventory_item_id, movement_date, movement_type,
            quantity, unit_cost, total_cost, stock_after, average_cost_after,
            reference, document_type, document_number, notes)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (item_id, movement_date, movement_type,
         quantity, unit_cost, total_cost, new_stock, new_avg,
         reference, doc_type, doc_number, notes)
    )

    execute_db(
        "UPDATE inventory_items SET current_stock=?, average_cost=?, updated_at=datetime('now') WHERE id=?",
        (new_stock, new_avg, item_id)
    )

    execute_db(
        "INSERT INTO audit_logs (user_id, company_id, entity_type, entity_id, action, new_value) VALUES (?,?,?,?,?,?)",
        (session['user_id'], company_id, 'kardex', item_id,
         f'movement_{movement_type}',
         f"{movement_type}: {abs(quantity):.3f} @ S/ {unit_cost:.4f}")
    )

    # Verificar stock mínimo y enviar alerta si es necesario
    _check_safety_stock(item_id, company_id, new_stock)

    bajo_minimo = (item['safety_stock'] and new_stock < float(item['safety_stock']))

    return jsonify({
        'success':    True,
        'new_stock':  new_stock,
        'new_avg':    round(new_avg, 4),
        'bajo_minimo': bajo_minimo,
        'message':    f'Movimiento registrado. Stock: {new_stock:.3f} | Costo prom: S/ {new_avg:.4f}'
                      + (' ⚠️ Stock bajo mínimo' if bajo_minimo else '')
    })


@kardex_bp.route('/kardex/item/<int:item_id>/delete', methods=['POST'])
@login_required
@company_required
def delete_item(item_id):
    company_id = session['company_id']
    execute_db(
        "UPDATE inventory_items SET is_active=0 WHERE id=? AND company_id=?",
        (item_id, company_id)
    )
    flash('Ítem desactivado.', 'success')
    return redirect(url_for('kardex.index'))


@kardex_bp.route('/kardex/export')
@login_required
@company_required
def export_kardex():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    company_id = session['company_id']
    company = query_db("SELECT * FROM companies WHERE id=?", (company_id,), one=True)
    items = query_db(
        """SELECT ii.*, u.code as unit_code FROM inventory_items ii
           LEFT JOIN units_of_measure u ON ii.unit_id=u.id
           WHERE ii.company_id=? AND ii.is_active=1 ORDER BY ii.name""",
        (company_id,)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Kardex"

    side   = Side(style='thin', color="CBD5E1")
    border = Border(left=side, right=side, top=side, bottom=side)

    ws.merge_cells('A1:H1')
    ws['A1'] = f"KARDEX DE INVENTARIO — {company['business_name']}"
    ws['A1'].font      = Font(bold=True, size=13, color="FFFFFF")
    ws['A1'].fill      = PatternFill("solid", fgColor="1F2937")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ["Código","Nombre","Categoría","Unidad","Stock Actual","Stock Mínimo","Costo Promedio","Valor Total"]
    widths  = [12, 32, 14, 10, 14, 12, 16, 16]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor="F97316")
        cell.alignment = Alignment(horizontal='center')
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    total_value = 0
    for row_idx, item in enumerate(items, 4):
        val = (item['current_stock'] or 0) * (item['average_cost'] or 0)
        total_value += val
        bajo = item['safety_stock'] and (item['current_stock'] or 0) < float(item['safety_stock'] or 0)
        fill_c = "FEE2E2" if bajo else ("FFF8F1" if row_idx % 2 == 0 else "FFFFFF")
        row_data = [
            item['code'], item['name'], item['category'],
            item['unit_code'] or '',
            item['current_stock'] or 0,
            item['safety_stock'] or 0,
            item['average_cost'] or 0,
            val,
        ]
        for col, v in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.fill   = PatternFill("solid", fgColor=fill_c)
            cell.border = border
            if col in (5, 6, 7, 8):
                cell.number_format = '"S/" #,##0.00'
                cell.alignment = Alignment(horizontal='right')

    last = len(items) + 4
    ws.cell(row=last, column=7, value="TOTAL").font = Font(bold=True)
    tot_cell = ws.cell(row=last, column=8, value=total_value)
    tot_cell.number_format = '"S/" #,##0.00'
    tot_cell.font = Font(bold=True)
    tot_cell.fill = PatternFill("solid", fgColor="FED7AA")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"kardex_{date.today()}.xlsx"
    )
