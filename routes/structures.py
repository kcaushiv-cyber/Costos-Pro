from flask import Blueprint, render_template, session, redirect, url_for, request, flash
from models.database import query_db, execute_db
from routes.dashboard import login_required, company_required, module_required

structures_bp = Blueprint('structures', __name__)


@structures_bp.route('/structures')
@login_required
@company_required
@module_required('structures', write=False)
def index():
    session['active_module'] = 'structures'
    company_id = session['company_id']
    structures = query_db(
        """SELECT cs.*, p.name as period_name, ps.name as product_name,
                  COALESCE((
                    SELECT SUM(ci.total_cost)
                    FROM cost_items ci
                    JOIN cost_categories cc ON ci.category_id = cc.id
                    WHERE cc.structure_id = cs.id
                  ), 0) as unit_cost_total
           FROM cost_structures cs
           LEFT JOIN periods p ON cs.period_id = p.id
           LEFT JOIN products_services ps ON cs.product_id = ps.id
           WHERE cs.company_id = ?
           ORDER BY cs.created_at DESC""",
        (company_id,)
    )
    return render_template('structures/index.html', structures=structures)


@structures_bp.route('/structures/new', methods=['GET', 'POST'])
@login_required
@company_required
@module_required('structures', write=True)
def new():
    company_id = session['company_id']
    period_id = session.get('period_id')

    if request.method == 'POST':
        d = request.form
        sid = execute_db(
            """INSERT INTO cost_structures
               (company_id, period_id, name, product_id, product_service,
                structure_date, monthly_production, sale_price, igv_rate,
                desired_margin, origin, image_path)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (company_id, period_id,
             d['name'],
             d.get('product_id') or None,
             d.get('product_service'),
             d.get('structure_date'),
             float(d.get('monthly_production', 0)),
             float(d.get('sale_price', 0)),
             float(d.get('igv_rate', 0.18)),
             float(d.get('desired_margin', 0)),
             'manual',
             d.get('image_path') or None)
        )
        # Crear categorías básicas MP, MOD, CIF
        for cat_name, cat_type in [
            ('Materia Prima (MP)', 'mp'),
            ('Mano de Obra Directa (MOD)', 'mod'),
            ('Costos Indirectos de Fabricación (CIF)', 'cif'),
            ('Gastos Administrativos', 'gasto_admin'),
            ('Gastos de Ventas', 'gasto_ventas'),
        ]:
            execute_db(
                "INSERT INTO cost_categories (structure_id, name, category_type) VALUES (?,?,?)",
                (sid, cat_name, cat_type)
            )
        flash('Estructura de costos creada', 'success')
        return redirect(url_for('structures.detail', sid=sid))

    products = query_db(
        "SELECT id, code, name FROM products_services WHERE company_id=? AND is_active=1 ORDER BY name",
        (company_id,)
    )
    return render_template('structures/new.html', products=products)


@structures_bp.route('/structures/<int:sid>')
@login_required
@company_required
@module_required('structures', write=False)
def detail(sid):
    company_id = session['company_id']
    structure = query_db(
        "SELECT cs.*, p.name as period_name FROM cost_structures cs LEFT JOIN periods p ON cs.period_id = p.id WHERE cs.id=? AND cs.company_id=?",
        (sid, company_id), one=True
    )
    if not structure:
        return redirect(url_for('structures.index'))

    categories = query_db(
        "SELECT * FROM cost_categories WHERE structure_id=? ORDER BY order_index, category_type",
        (sid,)
    )
    items = query_db(
        """SELECT ci.*, cc.name as category_name
           FROM cost_items ci
           JOIN cost_categories cc ON ci.category_id = cc.id
           WHERE cc.structure_id = ?
           ORDER BY cc.order_index, ci.order_index""",
        (sid,)
    )
    return render_template('structures/detail.html',
                           structure=structure, categories=categories, items=items)


@structures_bp.route('/structures/<int:sid>/items/new', methods=['POST'])
@login_required
@company_required
@module_required('structures', write=True)
def new_item(sid):
    d = request.form
    category_id = d.get('category_id')
    qty = float(d.get('quantity', 0))
    unit_cost = float(d.get('unit_cost', 0))
    execute_db(
        """INSERT INTO cost_items
           (category_id, description, cost_type, unit, quantity, unit_cost, total_cost)
           VALUES (?,?,?,?,?,?,?)""",
        (category_id, d['description'], d.get('cost_type', 'variable'),
         d.get('unit'), qty, unit_cost, qty * unit_cost)
    )
    flash('Ítem agregado', 'success')
    return redirect(url_for('structures.detail', sid=sid))


@structures_bp.route('/structures/<int:sid>/items/<int:iid>/delete', methods=['POST'])
@login_required
@company_required
@module_required('structures', write=True)
def delete_item(sid, iid):
    execute_db("DELETE FROM cost_items WHERE id=?", (iid,))
    return redirect(url_for('structures.detail', sid=sid))


@structures_bp.route('/structures/<int:sid>/edit', methods=['GET', 'POST'])
@login_required
@company_required
@module_required('structures', write=True)
def edit(sid):
    company_id = session['company_id']
    structure = query_db("SELECT * FROM cost_structures WHERE id=? AND company_id=?",
                         (sid, company_id), one=True)
    if not structure:
        flash('Estructura no encontrada', 'error')
        return redirect(url_for('structures.index'))

    if request.method == 'POST':
        d = request.form
        execute_db(
            """UPDATE cost_structures SET
               name=?, product_service=?, sale_price=?,
               igv_rate=?, desired_margin=?, monthly_production=?,
               image_path=COALESCE(NULLIF(?, ''), image_path)
               WHERE id=? AND company_id=?""",
            (d.get('name'), d.get('product_service'),
             float(d.get('sale_price', 0)),
             float(d.get('igv_rate', 0.18)),
             float(d.get('desired_margin', 0)),
             float(d.get('monthly_production', 0)),
             d.get('image_path') or None,
             sid, company_id)
        )
        flash('Estructura actualizada', 'success')
        return redirect(url_for('structures.detail', sid=sid))

    products = query_db("SELECT id, code, name FROM products_services WHERE company_id=? AND is_active=1", (company_id,))
    return render_template('structures/edit.html', structure=structure, products=products)


@structures_bp.route('/structures/<int:sid>/delete', methods=['POST'])
@login_required
@company_required
@module_required('structures', write=True)
def delete(sid):
    company_id = session['company_id']
    execute_db("DELETE FROM cost_structures WHERE id=? AND company_id=?", (sid, company_id))
    flash('Estructura eliminada', 'success')
    return redirect(url_for('structures.index'))


@structures_bp.route('/structures/<int:sid>/export')
@login_required
@company_required
@module_required('structures', write=False)
def export_excel(sid):
    """Exporta la estructura de costos a Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file
    import io

    company_id = session['company_id']
    structure  = query_db("SELECT * FROM cost_structures WHERE id=? AND company_id=?",
                          (sid, company_id), one=True)
    if not structure:
        flash('Estructura no encontrada', 'error')
        return redirect(url_for('structures.index'))

    cats  = query_db("SELECT * FROM cost_categories WHERE structure_id=? ORDER BY order_index", (sid,))
    items = query_db(
        """SELECT ci.*, cc.name as cat_name, cc.category_type
           FROM cost_items ci JOIN cost_categories cc ON ci.category_id=cc.id
           WHERE cc.structure_id=? ORDER BY cc.order_index, ci.order_index""",
        (sid,)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Estructura de Costos'

    fill_h   = PatternFill('solid', fgColor='1F2937')
    fill_cat = PatternFill('solid', fgColor='F97316')
    fill_alt = PatternFill('solid', fgColor='FFF8F1')
    font_w   = Font(bold=True, color='FFFFFF', size=10)
    font_b   = Font(bold=True, size=10)
    side     = Side(style='thin', color='D1D5DB')
    border   = Border(left=side, right=side, top=side, bottom=side)

    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Estructura de Costos: {structure['name']}"
    ws['A1'].font  = Font(bold=True, color='FFFFFF', size=13)
    ws['A1'].fill  = fill_h
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws['A2'] = f"Empresa: {session.get('company_name','')}   |   Período: {session.get('period_name','')}   |   Precio venta: S/ {structure['sale_price'] or 0:,.2f}"
    ws['A2'].font = Font(italic=True, size=10, color='6B7280')
    ws.merge_cells('A2:G2')

    # Headers
    headers = ['Categoría','Descripción','Tipo','Unidad','Cantidad','Costo Unit. (S/)','Total (S/)']
    widths  = [18, 30, 14, 8, 10, 14, 14]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = font_w; cell.fill = fill_cat; cell.border = border
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 4
    total_general = 0
    for item in items:
        fill = fill_alt if row % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        vals = [item['cat_name'], item['description'], item['cost_type'],
                item['unit'] or '', item['quantity'] or '', item['unit_cost'] or 0, item['total_cost'] or 0]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.fill = fill; cell.border = border
            if col >= 5:
                cell.number_format = '"S/" #,##0.00'
                cell.alignment = Alignment(horizontal='right')
        total_general += float(item['total_cost'] or 0)
        row += 1

    # Fila total
    ws.cell(row=row, column=1, value='COSTO TOTAL').font = font_b
    ws.cell(row=row, column=1).fill = fill_h
    ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
    ws.merge_cells(f'A{row}:F{row}')
    total_cell = ws.cell(row=row, column=7, value=total_general)
    total_cell.number_format = '"S/" #,##0.00'
    total_cell.font = Font(bold=True, color='FFFFFF')
    total_cell.fill = fill_h
    total_cell.alignment = Alignment(horizontal='right')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f"estructura_{structure['name'].replace(' ','_')}.xlsx")
