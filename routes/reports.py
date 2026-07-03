from flask import Blueprint, render_template, session, redirect, url_for, request, jsonify, send_file, flash
from models.database import query_db
from routes.dashboard import login_required, company_required
from services.report_service import (
    get_executive_summary, get_cost_structure_report,
    get_abc_report, get_labor_report,
    get_inventory_report, get_quality_report
)
from services.claude_service import send_message
import io
from datetime import date

reports_bp = Blueprint('reports', __name__)


@reports_bp.route('/reports')
@login_required
@company_required
def index():
    session['active_module'] = 'reports'
    company_id = session['company_id']
    period_id  = session.get('period_id')
    summary    = get_executive_summary(company_id, period_id)
    return render_template('reports/index.html', summary=summary)


# ── Reporte 1: Ejecutivo ───────────────────────────────────────────────────

@reports_bp.route('/reports/executive')
@login_required
@company_required
def executive():
    company_id = session['company_id']
    period_id  = session.get('period_id')
    summary    = get_executive_summary(company_id, period_id)
    return render_template('reports/executive.html', summary=summary)


# ── Reporte 2: Estructura de Costos ───────────────────────────────────────

@reports_bp.route('/reports/cost-structures')
@login_required
@company_required
def cost_structures():
    company_id = session['company_id']
    period_id  = session.get('period_id')
    data       = get_cost_structure_report(company_id, period_id)
    return render_template('reports/cost_structures.html', data=data)


# ── Reporte 3: ABC ────────────────────────────────────────────────────────

@reports_bp.route('/reports/abc')
@login_required
@company_required
def abc():
    company_id = session['company_id']
    period_id  = session.get('period_id')
    data       = get_abc_report(company_id, period_id)
    return render_template('reports/abc.html', data=data)


# ── Reporte 4: Personal ───────────────────────────────────────────────────

@reports_bp.route('/reports/labor')
@login_required
@company_required
def labor():
    company_id = session['company_id']
    data       = get_labor_report(company_id)
    return render_template('reports/labor.html', data=data)


# ── Reporte 5: Inventario ─────────────────────────────────────────────────

@reports_bp.route('/reports/inventory')
@login_required
@company_required
def inventory():
    company_id = session['company_id']
    data       = get_inventory_report(company_id)
    return render_template('reports/inventory.html', data=data)


# ── Reporte 6: Calidad ────────────────────────────────────────────────────

@reports_bp.route('/reports/quality')
@login_required
@company_required
def quality():
    company_id = session['company_id']
    period_id  = session.get('period_id')
    data       = get_quality_report(company_id, period_id)
    return render_template('reports/quality.html', data=data)


# ── Narrativa gerencial con IA ─────────────────────────────────────────────

@reports_bp.route('/reports/ai-narrative', methods=['POST'])
@login_required
@company_required
def ai_narrative():
    """Genera narrativa ejecutiva usando el agente 'reportes'."""
    company_id = session['company_id']
    period_id  = session.get('period_id')
    report_type = request.get_json().get('report_type', 'executive')

    summary = get_executive_summary(company_id, period_id)

    prompt = (
        f"Genera un reporte gerencial ejecutivo para {session.get('company_name')} "
        f"sector {session.get('sector')}, período {session.get('period_name')}.\n\n"
        f"Datos clave:\n"
        f"- Ventas proyectadas: S/ {summary['presupuesto']['ventas']:,.2f}\n"
        f"- Utilidad bruta: S/ {summary['presupuesto']['utilidad']:,.2f} "
        f"({summary['presupuesto']['margen_pct']}%)\n"
        f"- Personal activo: {summary['personal']['total']} personas, "
        f"costo mensual S/ {summary['personal']['costo_mes']:,.2f}\n"
        f"- Stock valorizado: S/ {summary['inventario']['valor']:,.2f} "
        f"({summary['inventario']['bajo_minimo']} ítems bajo mínimo)\n"
        f"- Costo de calidad: S/ {summary['calidad']['total']:,.2f} "
        f"({summary['calidad']['pct_fallas']}% en fallas)\n\n"
        f"Analiza los KPIs, identifica riesgos y oportunidades, y da 3 recomendaciones concretas."
    )

    result = send_message(prompt, agent_type='reportes')
    return jsonify(result)


# ── Exportador Excel ───────────────────────────────────────────────────────

@reports_bp.route('/reports/export/<string:report_type>')
@login_required
@company_required
def export_excel(report_type):
    """Exporta cualquier reporte a Excel con formato corporativo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    company_id   = session['company_id']
    period_id    = session.get('period_id')
    company_name = session.get('company_name', '')
    period_name  = session.get('period_name', '')

    wb = Workbook()

    # Estilos compartidos
    side    = Side(style='thin', color='D1D5DB')
    border  = Border(left=side, right=side, top=side, bottom=side)
    fill_header = PatternFill('solid', fgColor='1F2937')
    fill_sub    = PatternFill('solid', fgColor='F97316')
    fill_alt    = PatternFill('solid', fgColor='FFF8F1')
    font_h  = Font(bold=True, color='FFFFFF', size=11)
    font_sub= Font(bold=True, color='FFFFFF', size=10)
    font_b  = Font(bold=True, size=10)
    align_c = Alignment(horizontal='center', vertical='center')
    align_r = Alignment(horizontal='right')

    def write_title(ws, title, cols):
        ws.merge_cells(f'A1:{get_column_letter(cols)}1')
        ws['A1'] = f"{title} — {company_name} | {period_name} | {date.today()}"
        ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
        ws['A1'].fill = fill_header
        ws['A1'].alignment = align_c
        ws.row_dimensions[1].height = 26

    def write_headers(ws, row, headers, widths):
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = font_sub; c.fill = fill_sub
            c.alignment = align_c; c.border = border
            ws.column_dimensions[get_column_letter(col)].width = w

    def write_row(ws, row_idx, values, number_cols=()):
        fill = fill_alt if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.fill = fill; c.border = border
            if col in number_cols:
                c.number_format = '"S/" #,##0.00'
                c.alignment = align_r

    # ── Resumen Ejecutivo ──────────────────────────────────────────────────
    if report_type == 'executive':
        ws = wb.active; ws.title = 'Resumen Ejecutivo'
        summary = get_executive_summary(company_id, period_id)
        write_title(ws, 'REPORTE EJECUTIVO', 3)

        sections = [
            ('PRESUPUESTO', [
                ('Ventas proyectadas', summary['presupuesto']['ventas']),
                ('Costo producción',   summary['presupuesto']['costo']),
                ('Utilidad bruta',     summary['presupuesto']['utilidad']),
                ('Margen bruto %',     summary['presupuesto']['margen_pct']),
            ]),
            ('PERSONAL', [
                ('Total trabajadores',  summary['personal']['total']),
                ('Costo planilla/mes',  summary['personal']['costo_mes']),
                ('Costo MOD/mes',       summary['personal']['costo_mod']),
                ('Costo MOI/mes',       summary['personal']['costo_moi']),
            ]),
            ('INVENTARIO', [
                ('Ítems activos',       summary['inventario']['total_items']),
                ('Stock valorizado',    summary['inventario']['valor']),
                ('Ítems bajo mínimo',   summary['inventario']['bajo_minimo']),
            ]),
            ('CALIDAD (PAF)', [
                ('Costo calidad total', summary['calidad']['total']),
                ('Costos de fallas',    summary['calidad']['fallas']),
                ('Prevención',          summary['calidad']['prevencion']),
                ('% sobre ventas',      summary['calidad']['pct_ventas']),
            ]),
        ]
        row = 3
        for section_name, items in sections:
            ws.cell(row=row, column=1, value=section_name).font = font_b
            ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='E5D5C5')
            row += 1
            for label, value in items:
                ws.cell(row=row, column=1, value=label)
                c = ws.cell(row=row, column=2, value=value)
                if isinstance(value, float) and value > 1:
                    c.number_format = '"S/" #,##0.00'
                    c.alignment = align_r
                row += 1
            row += 1
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 18

    # ── Personal ───────────────────────────────────────────────────────────
    elif report_type == 'labor':
        ws = wb.active; ws.title = 'Personal y MOD'
        data = get_labor_report(company_id)
        headers = ['Código','Nombre','Depto','Cargo','Tipo','Sueldo básico','Costo total/mes','S/./hora','Hrs disponibles']
        widths  = [10,28,18,20,8,14,16,10,14]
        write_title(ws, 'REPORTE DE PERSONAL Y MOD', len(headers))
        write_headers(ws, 3, headers, widths)
        for i, e in enumerate(data['employees'], 4):
            write_row(ws, i, [
                e['code'], e['full_name'], e.get('dept_name',''), e.get('position_name',''),
                e['labor_type'], e['basic_salary'], e['total_monthly_cost'],
                e['cost_per_hour'], e['available_hours_month']
            ], number_cols=(6,7,8))

    # ── Inventario ─────────────────────────────────────────────────────────
    elif report_type == 'inventory':
        ws = wb.active; ws.title = 'Inventario'
        data = get_inventory_report(company_id)
        headers = ['Código','Nombre','Categoría','Unidad','Stock actual','Costo prom.','Valor total','Stock mín.','Estado']
        widths  = [12,28,12,8,12,14,14,10,10]
        write_title(ws, 'REPORTE DE INVENTARIO', len(headers))
        write_headers(ws, 3, headers, widths)
        for i, item in enumerate(data['item_list'], 4):
            estado = '⚠ Bajo' if item.get('safety_stock') and float(item['current_stock'] or 0) < float(item['safety_stock'] or 0) else 'OK'
            write_row(ws, i, [
                item['code'], item['name'], item['category'],
                item.get('unit_code',''), item['current_stock'],
                item['average_cost'], item['valor_total'],
                item['safety_stock'] or 0, estado
            ], number_cols=(6,7))

    # ── ABC ────────────────────────────────────────────────────────────────
    elif report_type == 'abc':
        ws = wb.active; ws.title = 'Resultados ABC'
        data = get_abc_report(company_id, period_id)
        headers = ['Modelo','Código','Objeto de Costo','Qty/mes','C. Producción','Costo Total','P. Venta','CU ABC','CU Trad.','Diferencia']
        widths  = [20,10,24,8,14,14,14,12,12,12]
        write_title(ws, 'REPORTE COSTEO ABC', len(headers))
        write_headers(ws, 3, headers, widths)
        for i, r in enumerate(data['results'], 4):
            write_row(ws, i, [
                r['model'], r['obj_code'], r['obj_name'], r['qty'],
                r['prod_cost'], r['total_cost'], r['sale_price'],
                r['unit_cost_abc'], r['unit_cost_trad'], r['diff']
            ], number_cols=(5,6,7,8,9,10))

    # ── Calidad ────────────────────────────────────────────────────────────
    elif report_type == 'quality':
        ws = wb.active; ws.title = 'Costos de Calidad'
        data = get_quality_report(company_id, period_id)
        # Hoja 1: Detalle
        headers = ['Actividad','Categoría','Responsable','Costo mensual','Costo anual','Clasificado por IA']
        widths  = [32,16,20,14,14,16]
        write_title(ws, 'REPORTE COSTOS DE CALIDAD PAF', len(headers))
        write_headers(ws, 3, headers, widths)
        for i, c in enumerate(data['costs'], 4):
            write_row(ws, i, [
                c['activity_name'], c['category'].replace('_',' ').title(),
                c.get('responsible',''), c['monthly_cost'], c['annual_cost'],
                'Sí' if c.get('ai_classified') else 'No'
            ], number_cols=(4,5))
        # Hoja 2: Resumen PAF
        ws2 = wb.create_sheet('Resumen PAF')
        ws2['A1'] = 'Categoría';  ws2['B1'] = 'Mensual';  ws2['C1'] = '% Total'
        for cell in [ws2['A1'], ws2['B1'], ws2['C1']]:
            cell.font = font_sub; cell.fill = fill_sub
        for row, (cat, val) in enumerate([
            ('Prevención',    data['totals']['prevencion']),
            ('Evaluación',    data['totals']['evaluacion']),
            ('Falla Interna', data['totals']['falla_interna']),
            ('Falla Externa', data['totals']['falla_externa']),
        ], 2):
            ws2.cell(row=row, column=1, value=cat)
            ws2.cell(row=row, column=2, value=val).number_format = '"S/" #,##0.00'
            pct = round(val / data['grand_total'] * 100, 1) if data['grand_total'] else 0
            ws2.cell(row=row, column=3, value=pct).number_format = '0.0"%"'
        ws2.cell(row=6, column=1, value='TOTAL').font = font_b
        ws2.cell(row=6, column=2, value=data['grand_total']).number_format = '"S/" #,##0.00'

    else:
        flash('Tipo de reporte no válido', 'error')
        return redirect(url_for('reports.index'))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"reporte_{report_type}_{session.get('company_name','').replace(' ','_')}_{date.today()}.xlsx"
    )


# ── Historial de cambios ───────────────────────────────────────────────────

@reports_bp.route('/reports/history')
@login_required
@company_required
def history():
    company_id = session['company_id']
    page = request.args.get('page', 1, type=int)
    limit = 50
    offset = (page - 1) * limit

    logs = query_db(
        """SELECT al.*, u.full_name as user_name
           FROM audit_logs al
           JOIN users u ON al.user_id = u.id
           WHERE al.company_id=?
           ORDER BY al.created_at DESC
           LIMIT ? OFFSET ?""",
        (company_id, limit, offset)
    )
    total = query_db(
        "SELECT COUNT(*) as n FROM audit_logs WHERE company_id=?",
        (company_id,), one=True
    )
    return render_template('reports/history.html',
                           logs=logs,
                           total=total['n'] if total else 0,
                           page=page, limit=limit)


# ── Exportador PDF ─────────────────────────────────────────────────────────

@reports_bp.route('/reports/export/pdf/<string:report_type>')
@login_required
@company_required
def export_pdf(report_type):
    """Genera PDF usando HTML→weasyprint o fallback a HTML imprimible."""
    company_id   = session['company_id']
    period_id    = session.get('period_id')
    company_name = session.get('company_name', '')
    period_name  = session.get('period_name', '')

    # Intentar weasyprint, si no está disponible generar HTML imprimible
    try:
        import weasyprint
        has_weasyprint = True
    except ImportError:
        has_weasyprint = False

    if report_type == 'executive':
        from services.report_service import get_executive_summary
        data = get_executive_summary(company_id, period_id)
        html_content = render_template('reports/pdf/executive.html',
                                       summary=data,
                                       company_name=company_name,
                                       period_name=period_name)
    elif report_type == 'abc':
        from services.report_service import get_abc_report
        data = get_abc_report(company_id, period_id)
        html_content = render_template('reports/pdf/abc.html',
                                       data=data,
                                       company_name=company_name,
                                       period_name=period_name)
    else:
        flash('Tipo de reporte PDF no disponible', 'error')
        return redirect(url_for('reports.index'))

    if has_weasyprint:
        pdf = weasyprint.HTML(string=html_content).write_pdf()
        from flask import Response
        return Response(pdf,
                        mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment;filename=reporte_{report_type}_{period_name}.pdf'})
    else:
        # Devolver HTML para imprimir
        from flask import Response
        return Response(
            html_content + '<script>window.onload=function(){window.print()}</script>',
            mimetype='text/html'
        )
