from flask import Blueprint, render_template, session, request, jsonify, flash, redirect, url_for
from models.database import query_db, execute_db
from routes.dashboard import login_required, company_required, module_required
import io

import_bp = Blueprint('import_data', __name__)


@import_bp.route('/import')
@login_required
@company_required
@module_required('import_data', write=False)
def index():
    session['active_module'] = 'import_data'
    company_id = session['company_id']
    structures = query_db("SELECT id, name FROM cost_structures WHERE company_id=? ORDER BY created_at DESC", (company_id,))
    abc_models = query_db("SELECT id, name FROM abc_models WHERE company_id=? ORDER BY created_at DESC", (company_id,))
    return render_template('import_data/index.html', structures=structures, abc_models=abc_models)


@import_bp.route('/import/template/<string:entity>')
@login_required
@module_required('import_data', write=False)
def download_template(entity):
    """Descarga plantilla Excel para importación."""
    import openpyxl
    from flask import send_file
    templates = {
        'employees': (['codigo','nombre','sueldo_basico','tipo_mod','horas_disponibles'],
                      [['EMP001','Juan Pérez García',1200,'MOD',192],['EMP002','María López',1500,'admin',192]]),
        'products':  (['codigo','nombre','categoria','precio_venta','costo_estandar'],
                      [['PROD001','Camisa M','producto',80,45],['SERV001','Diseño','servicio',150,0]]),
        'resources': (['codigo','nombre','categoria','monto_mensual'],
                      [['REC001','Energía eléctrica','energia',1200],['REC002','Alquiler','servicios',3000]]),
        'inventory': (['codigo','nombre','categoria','stock_inicial','costo_unitario'],
                      [['INS001','Tela algodón','insumo',500,8.5],['INS002','Hilo','insumo',1000,0.5]]),
    }
    if entity not in templates:
        from flask import abort
        abort(404)
    headers, rows = templates[entity]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = entity
    ws.append(headers)
    for row in rows:
        ws.append(row)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'plantilla_{entity}.xlsx')


@import_bp.route('/import/excel', methods=['POST'])
@login_required
@company_required
@module_required('import_data', write=True)
def import_excel():
    """Importa datos desde Excel. Detecta el tipo por las columnas."""
    import openpyxl

    company_id = session['company_id']
    period_id  = session.get('period_id')
    entity     = request.form.get('entity', 'employees')

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No se envió archivo'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Solo se aceptan archivos .xlsx'}), 400

    try:
        wb  = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws  = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({'success': False, 'error': 'Archivo vacío'}), 400

        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        data_rows = rows[1:]
        inserted = 0
        errors   = []

        if entity == 'employees':
            # Columnas esperadas: codigo, nombre, sueldo_basico, tipo_mod, horas_disponibles
            for i, row in enumerate(data_rows, 2):
                if not any(row):
                    continue
                try:
                    d = dict(zip(headers, row))
                    basic = float(d.get('sueldo_basico') or d.get('basic_salary') or 0)
                    grat  = basic / 6
                    cts   = basic * 7 / 72
                    ess   = basic * 0.09
                    total = basic + grat + cts + ess
                    hours = float(d.get('horas_disponibles') or d.get('available_hours') or 192)
                    execute_db(
                        """INSERT OR IGNORE INTO employees
                           (company_id, code, full_name, basic_salary,
                            gratification_monthly, cts_monthly, essalud,
                            total_monthly_cost, cost_per_hour,
                            available_hours_month, labor_type, is_active, in_payroll)
                           VALUES (?,?,?,?,?,?,?,?,?,?,?,1,1)""",
                        (company_id,
                         str(d.get('codigo') or d.get('code') or f'EMP{i:03d}'),
                         str(d.get('nombre') or d.get('full_name') or ''),
                         basic, round(grat, 2), round(cts, 2), round(ess, 2),
                         round(total, 2), round(total / hours, 4) if hours else 0,
                         hours,
                         str(d.get('tipo_mod') or d.get('labor_type') or 'MOD'))
                    )
                    inserted += 1
                except Exception as e:
                    errors.append(f'Fila {i}: {str(e)[:50]}')

        elif entity == 'products':
            for i, row in enumerate(data_rows, 2):
                if not any(row):
                    continue
                try:
                    d = dict(zip(headers, row))
                    execute_db(
                        """INSERT OR IGNORE INTO products_services
                           (company_id, code, name, category, sale_price, standard_cost)
                           VALUES (?,?,?,?,?,?)""",
                        (company_id,
                         str(d.get('codigo') or d.get('code') or f'PROD{i:03d}'),
                         str(d.get('nombre') or d.get('name') or ''),
                         str(d.get('categoria') or d.get('category') or 'producto'),
                         float(d.get('precio_venta') or d.get('sale_price') or 0),
                         float(d.get('costo_estandar') or d.get('standard_cost') or 0))
                    )
                    inserted += 1
                except Exception as e:
                    errors.append(f'Fila {i}: {str(e)[:50]}')

        elif entity == 'resources':
            for i, row in enumerate(data_rows, 2):
                if not any(row):
                    continue
                try:
                    d = dict(zip(headers, row))
                    monthly = float(d.get('monto_mensual') or d.get('monthly_amount') or 0)
                    execute_db(
                        """INSERT OR IGNORE INTO resources
                           (company_id, period_id, code, name, category, monthly_amount, annual_amount)
                           VALUES (?,?,?,?,?,?,?)""",
                        (company_id, period_id,
                         str(d.get('codigo') or d.get('code') or f'REC{i:03d}'),
                         str(d.get('nombre') or d.get('name') or ''),
                         str(d.get('categoria') or d.get('category') or 'otros'),
                         monthly, monthly * 12)
                    )
                    inserted += 1
                except Exception as e:
                    errors.append(f'Fila {i}: {str(e)[:50]}')

        elif entity == 'inventory':
            from datetime import date
            for i, row in enumerate(data_rows, 2):
                if not any(row):
                    continue
                try:
                    d = dict(zip(headers, row))
                    stock = float(d.get('stock_inicial') or d.get('initial_stock') or 0)
                    cost  = float(d.get('costo_unitario') or d.get('unit_cost') or 0)
                    iid = execute_db(
                        """INSERT OR IGNORE INTO inventory_items
                           (company_id, code, name, category, current_stock, average_cost, valuation_method, is_active)
                           VALUES (?,?,?,?,?,?,?,1)""",
                        (company_id,
                         str(d.get('codigo') or d.get('code') or f'INV{i:03d}'),
                         str(d.get('nombre') or d.get('name') or ''),
                         str(d.get('categoria') or d.get('category') or 'insumo'),
                         stock, cost, 'promedio')
                    )
                    if iid and stock > 0:
                        execute_db(
                            """INSERT INTO kardex_movements
                               (inventory_item_id, movement_date, movement_type,
                                quantity, unit_cost, total_cost, stock_after, average_cost_after, reference)
                               VALUES (?,?,?,?,?,?,?,?,?)""",
                            (iid, str(date.today()), 'entrada',
                             stock, cost, stock * cost, stock, cost, 'Saldo inicial importado')
                        )
                    inserted += 1
                except Exception as e:
                    errors.append(f'Fila {i}: {str(e)[:50]}')

        return jsonify({
            'success': True,
            'inserted': inserted,
            'total':    len(data_rows),
            'errors':   errors[:5],
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ── Duplicar registros ─────────────────────────────────────────────────────

@import_bp.route('/duplicate/structure/<int:sid>', methods=['POST'])
@login_required
@company_required
@module_required('import_data', write=True)
def duplicate_structure(sid):
    """Duplica una estructura de costos a otro período."""
    company_id = session['company_id']
    period_id  = session.get('period_id')

    struct = query_db("SELECT * FROM cost_structures WHERE id=? AND company_id=?",
                      (sid, company_id), one=True)
    if not struct:
        return jsonify({'success': False, 'error': 'Estructura no encontrada'}), 404

    # Crear copia
    new_id = execute_db(
        """INSERT INTO cost_structures
           (company_id, period_id, name, product_id, product_service,
            sale_price, igv_rate, desired_margin, origin, image_path)
           VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (company_id, period_id,
         f"Copia de {struct['name']}",
         struct['product_id'], struct['product_service'],
         struct['sale_price'], struct['igv_rate'],
         struct['desired_margin'], 'duplicado',
         struct['image_path'])
    )

    # Duplicar categorías e ítems
    cats = query_db("SELECT * FROM cost_categories WHERE structure_id=? ORDER BY order_index", (sid,))
    for cat in cats:
        cat_id = execute_db(
            "INSERT INTO cost_categories (structure_id, name, category_type, order_index) VALUES (?,?,?,?)",
            (new_id, cat['name'], cat['category_type'], cat['order_index'])
        )
        items = query_db("SELECT * FROM cost_items WHERE category_id=?", (cat['id'],))
        for item in items:
            execute_db(
                """INSERT INTO cost_items
                   (category_id, description, cost_type, unit, quantity, unit_cost, total_cost, order_index)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (cat_id, item['description'], item['cost_type'],
                 item['unit'], item['quantity'], item['unit_cost'],
                 item['total_cost'], item['order_index'])
            )

    return jsonify({'success': True, 'new_id': new_id,
                    'redirect': url_for('structures.detail', sid=new_id)})


@import_bp.route('/duplicate/abc/<int:mid>', methods=['POST'])
@login_required
@company_required
@module_required('import_data', write=True)
def duplicate_abc(mid):
    """Duplica un modelo ABC al período activo."""
    company_id = session['company_id']
    period_id  = session.get('period_id')

    model = query_db("SELECT * FROM abc_models WHERE id=? AND company_id=?",
                     (mid, company_id), one=True)
    if not model:
        return jsonify({'success': False, 'error': 'Modelo no encontrado'}), 404

    new_id = execute_db(
        """INSERT INTO abc_models
           (company_id, period_id, name, sector,
            admin_expense, sales_expense, financial_expense, desired_margin, image_path)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (company_id, period_id,
         f"Copia de {model['name']}",
         model['sector'],
         model['admin_expense'], model['sales_expense'],
         model['financial_expense'], model['desired_margin'],
         model['image_path'])
    )

    return jsonify({'success': True, 'new_id': new_id,
                    'redirect': url_for('abc.step1_resources', mid=new_id)})
