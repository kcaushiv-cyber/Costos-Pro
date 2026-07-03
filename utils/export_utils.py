import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side, 
                               numbers, GradientFill)
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                  Spacer, HRFlowable)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


def export_structure_excel(structure, company, categories_with_items, totals):
    """Exportar estructura de costos a Excel con formato profesional."""
    wb = Workbook()
    
    # ========== HOJA 1: ESTRUCTURA DE COSTOS ==========
    ws = wb.active
    ws.title = "Estructura de Costos"
    
    # Colores
    COLOR_DARK_BLUE = "1E3A5F"
    COLOR_BLUE = "2563EB"
    COLOR_LIGHT_BLUE = "DBEAFE"
    COLOR_CATEGORY = "1E40AF"
    COLOR_CATEGORY_BG = "EFF6FF"
    COLOR_ALT = "F8FAFC"
    COLOR_TOTAL = "FEF3C7"
    COLOR_GRAND_TOTAL = "FCD34D"
    
    def apply_border(cell, style='thin'):
        side = Side(style=style, color="CBD5E1")
        cell.border = Border(left=side, right=side, top=side, bottom=side)
    
    # --- ENCABEZADO ---
    ws.merge_cells('A1:I1')
    ws['A1'] = "ESTRUCTURA DE COSTOS EMPRESARIALES"
    ws['A1'].font = Font(name='Calibri', bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor=COLOR_DARK_BLUE)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells('A2:I2')
    ws['A2'] = f"{company['business_name']} | RUC: {company['ruc']} | Sector: {company['sector']}"
    ws['A2'].font = Font(name='Calibri', size=11, color="FFFFFF")
    ws['A2'].fill = PatternFill("solid", fgColor=COLOR_BLUE)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22
    
    ws.merge_cells('A3:I3')
    ws['A3'] = f"Estructura: {structure['name']} | Producto: {structure['product_service'] or 'N/A'} | Fecha: {datetime.now().strftime('%d/%m/%Y')}"
    ws['A3'].font = Font(name='Calibri', size=10, color=COLOR_DARK_BLUE)
    ws['A3'].fill = PatternFill("solid", fgColor=COLOR_LIGHT_BLUE)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 18
    
    ws.row_dimensions[4].height = 8
    
    # --- ENCABEZADOS DE TABLA ---
    headers = ["#", "Código", "Descripción", "Tipo", "Unidad", "Cantidad", "Costo Unit. (S/)", "Total (S/)", "% Part."]
    col_widths = [5, 10, 38, 10, 10, 10, 16, 14, 8]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(name='Calibri', bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLOR_CATEGORY)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[5].height = 24
    
    grand_total = totals.get('grand_total', 0)
    row_num = 6
    item_counter = 0
    
    for cat_data in categories_with_items:
        cat = cat_data['category']
        items = cat_data['items']
        cat_total = cat_data['total']
        
        # Fila de categoría
        ws.merge_cells(f'A{row_num}:B{row_num}')
        ws.merge_cells(f'C{row_num}:G{row_num}')
        ws.cell(row=row_num, column=1, value=f"▶  {cat['name'].upper()}")
        ws.cell(row=row_num, column=1).font = Font(name='Calibri', bold=True, size=11, color="FFFFFF")
        for col in range(1, 10):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = PatternFill("solid", fgColor=COLOR_CATEGORY)
            apply_border(cell)
        ws.cell(row=row_num, column=8, value=cat_total)
        ws.cell(row=row_num, column=8).font = Font(bold=True, color="FFFFFF", size=11)
        ws.cell(row=row_num, column=8).number_format = '"S/" #,##0.00'
        ws.cell(row=row_num, column=8).alignment = Alignment(horizontal='right')
        pct = (cat_total / grand_total * 100) if grand_total > 0 else 0
        ws.cell(row=row_num, column=9, value=f"{pct:.1f}%")
        ws.cell(row=row_num, column=9).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row_num, column=9).alignment = Alignment(horizontal='center')
        ws.row_dimensions[row_num].height = 22
        row_num += 1
        
        # Ítems
        for i, item in enumerate(items):
            item_counter += 1
            fill_color = "FFFFFF" if i % 2 == 0 else COLOR_ALT
            row_data = [
                item_counter,
                item.get('code', f"{item_counter}"),
                item['description'],
                item.get('cost_type', 'variable').capitalize(),
                item.get('unit', ''),
                item.get('quantity', 0),
                item.get('unit_cost', 0),
                item.get('total_cost', 0),
                f"{(item.get('total_cost', 0) / grand_total * 100):.1f}%" if grand_total > 0 else "0.0%"
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(name='Calibri', size=9)
                apply_border(cell)
                if col_idx in [6, 7, 8]:
                    cell.number_format = '"S/" #,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif col_idx in [1, 2, 4, 5, 9]:
                    cell.alignment = Alignment(horizontal='center')
            ws.row_dimensions[row_num].height = 16
            row_num += 1
        
        row_num += 1  # Espacio entre categorías
    
    # --- TOTAL GENERAL ---
    ws.merge_cells(f'A{row_num}:G{row_num}')
    ws.cell(row=row_num, column=1, value="COSTO TOTAL GENERAL")
    ws.cell(row=row_num, column=1).font = Font(name='Calibri', bold=True, size=13, color=COLOR_DARK_BLUE)
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=row_num, column=8, value=grand_total)
    ws.cell(row=row_num, column=8).font = Font(name='Calibri', bold=True, size=13, color=COLOR_DARK_BLUE)
    ws.cell(row=row_num, column=8).number_format = '"S/" #,##0.00'
    ws.cell(row=row_num, column=8).alignment = Alignment(horizontal='right')
    ws.cell(row=row_num, column=9, value="100.0%")
    ws.cell(row=row_num, column=9).font = Font(bold=True, color=COLOR_DARK_BLUE)
    ws.cell(row=row_num, column=9).alignment = Alignment(horizontal='center')
    for col in range(1, 10):
        ws.cell(row=row_num, column=col).fill = PatternFill("solid", fgColor=COLOR_GRAND_TOTAL)
        apply_border(ws.cell(row=row_num, column=col))
    ws.row_dimensions[row_num].height = 28
    row_num += 2
    
    # --- RESUMEN FINANCIERO ---
    monthly_prod = structure.get('monthly_production') or 0
    sale_price = structure.get('sale_price') or 0
    
    if monthly_prod > 0 or sale_price > 0:
        ws.merge_cells(f'A{row_num}:I{row_num}')
        ws.cell(row=row_num, column=1, value="RESUMEN FINANCIERO")
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=row_num, column=1).fill = PatternFill("solid", fgColor=COLOR_DARK_BLUE)
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
        row_num += 1
        
        resumen = []
        if monthly_prod > 0:
            costo_unit = grand_total / monthly_prod
            resumen.append(("Producción mensual estimada:", f"{monthly_prod:,.0f} unidades"))
            resumen.append(("Costo total mensual:", f"S/ {grand_total:,.2f}"))
            resumen.append(("Costo unitario:", f"S/ {costo_unit:,.4f}"))
        
        if sale_price > 0 and monthly_prod > 0:
            costo_unit = grand_total / monthly_prod
            precio_sin_igv = sale_price / 1.18
            utilidad = precio_sin_igv - costo_unit
            margen = (utilidad / precio_sin_igv * 100) if precio_sin_igv > 0 else 0
            resumen.append(("Precio de venta (con IGV):", f"S/ {sale_price:,.2f}"))
            resumen.append(("Precio sin IGV:", f"S/ {precio_sin_igv:,.2f}"))
            resumen.append(("Utilidad unitaria:", f"S/ {utilidad:,.4f}"))
            resumen.append(("Margen de utilidad:", f"{margen:.1f}%"))
        
        for label, value in resumen:
            ws.merge_cells(f'A{row_num}:F{row_num}')
            ws.cell(row=row_num, column=1, value=label)
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=10)
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='right')
            ws.merge_cells(f'G{row_num}:I{row_num}')
            ws.cell(row=row_num, column=7, value=value)
            ws.cell(row=row_num, column=7).font = Font(size=10, color=COLOR_BLUE)
            ws.cell(row=row_num, column=7).alignment = Alignment(horizontal='center')
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).fill = PatternFill("solid", fgColor=COLOR_LIGHT_BLUE)
                apply_border(ws.cell(row=row_num, column=col))
            row_num += 1
    
    # ========== HOJA 2: RESUMEN ==========
    ws2 = wb.create_sheet("Resumen por Categoría")
    
    ws2.merge_cells('A1:D1')
    ws2['A1'] = "RESUMEN POR CATEGORÍA"
    ws2['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws2['A1'].fill = PatternFill("solid", fgColor=COLOR_DARK_BLUE)
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 30
    
    ws2.merge_cells('A2:D2')
    ws2['A2'] = f"{company['business_name']} | {structure['name']}"
    ws2['A2'].fill = PatternFill("solid", fgColor=COLOR_LIGHT_BLUE)
    ws2['A2'].alignment = Alignment(horizontal='center')
    
    headers2 = ["Categoría", "Total (S/)", "% Participación", "# Ítems"]
    widths2 = [35, 15, 15, 10]
    for col_idx, (h, w) in enumerate(zip(headers2, widths2), 1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLOR_CATEGORY)
        cell.alignment = Alignment(horizontal='center')
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
        apply_border(cell)
    
    for i, cat_data in enumerate(categories_with_items, 5):
        cat = cat_data['category']
        cat_total = cat_data['total']
        pct = (cat_total / grand_total * 100) if grand_total > 0 else 0
        fill_c = COLOR_ALT if i % 2 == 0 else "FFFFFF"
        
        row_data = [cat['name'], cat_total, f"{pct:.1f}%", len(cat_data['items'])]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=i, column=col_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=fill_c)
            if col_idx == 2:
                cell.number_format = '"S/" #,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col_idx in [3, 4]:
                cell.alignment = Alignment(horizontal='center')
            apply_border(cell)
    
    ws2.cell(row=len(categories_with_items)+5, column=1, value="TOTAL GENERAL")
    ws2.cell(row=len(categories_with_items)+5, column=1).font = Font(bold=True)
    ws2.cell(row=len(categories_with_items)+5, column=2, value=grand_total)
    ws2.cell(row=len(categories_with_items)+5, column=2).number_format = '"S/" #,##0.00'
    ws2.cell(row=len(categories_with_items)+5, column=2).font = Font(bold=True)
    ws2.cell(row=len(categories_with_items)+5, column=3, value="100.0%")
    ws2.cell(row=len(categories_with_items)+5, column=3).font = Font(bold=True)
    for col in range(1, 5):
        ws2.cell(row=len(categories_with_items)+5, column=col).fill = PatternFill("solid", fgColor=COLOR_GRAND_TOTAL)
        apply_border(ws2.cell(row=len(categories_with_items)+5, column=col))
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_structure_pdf(structure, company, categories_with_items, totals):
    """Exportar estructura de costos a PDF con formato profesional usando ReportLab."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    styles = getSampleStyleSheet()
    COLOR_BLUE = colors.HexColor('#2563EB')
    COLOR_DARK = colors.HexColor('#1E3A5F')
    COLOR_LIGHT = colors.HexColor('#DBEAFE')
    COLOR_CAT = colors.HexColor('#1E40AF')
    COLOR_ALT = colors.HexColor('#F8FAFC')
    COLOR_YELLOW = colors.HexColor('#FCD34D')
    
    story = []
    
    # Título principal
    story.append(Paragraph(
        f"<font color='#1E3A5F'><b>ESTRUCTURA DE COSTOS EMPRESARIALES</b></font>",
        ParagraphStyle('Title', parent=styles['Title'], fontSize=18, alignment=TA_CENTER, spaceAfter=4)
    ))
    story.append(HRFlowable(width="100%", thickness=3, color=COLOR_DARK))
    story.append(Spacer(1, 6))
    
    # Info empresa
    story.append(Paragraph(
        f"<b>{company['business_name']}</b> &nbsp;|&nbsp; RUC: {company['ruc']} &nbsp;|&nbsp; Sector: {company['sector']}",
        ParagraphStyle('Empresa', parent=styles['Normal'], fontSize=11, alignment=TA_CENTER, textColor=COLOR_DARK)
    ))
    story.append(Paragraph(
        f"Estructura: <b>{structure['name']}</b> &nbsp;|&nbsp; Producto: {structure.get('product_service','N/A')} &nbsp;|&nbsp; Fecha: {datetime.now().strftime('%d/%m/%Y')}",
        ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, spaceAfter=12)
    ))
    
    grand_total = totals.get('grand_total', 0)
    
    # Tabla de costos
    col_widths_pdf = [1.2*cm, 2.2*cm, 8*cm, 2*cm, 2*cm, 2.2*cm, 2.8*cm, 2.8*cm, 1.8*cm]
    header_row = [
        Paragraph('<b>#</b>', ParagraphStyle('H', alignment=TA_CENTER)),
        Paragraph('<b>Código</b>', ParagraphStyle('H', alignment=TA_CENTER)),
        Paragraph('<b>Descripción</b>', ParagraphStyle('H', alignment=TA_CENTER)),
        Paragraph('<b>Tipo</b>', ParagraphStyle('H', alignment=TA_CENTER)),
        Paragraph('<b>Unidad</b>', ParagraphStyle('H', alignment=TA_CENTER)),
        Paragraph('<b>Cantidad</b>', ParagraphStyle('H', alignment=TA_CENTER)),
        Paragraph('<b>Costo Unit.</b>', ParagraphStyle('H', alignment=TA_CENTER)),
        Paragraph('<b>Total (S/)</b>', ParagraphStyle('H', alignment=TA_CENTER)),
        Paragraph('<b>%</b>', ParagraphStyle('H', alignment=TA_CENTER)),
    ]
    
    table_data = [header_row]
    table_styles_list = [
        ('BACKGROUND', (0, 0), (-1, 0), COLOR_CAT),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_ALT]),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
    ]
    
    item_counter = 0
    cat_row_indices = []
    
    for cat_data in categories_with_items:
        cat = cat_data['category']
        items = cat_data['items']
        cat_total = cat_data['total']
        pct_cat = (cat_total / grand_total * 100) if grand_total > 0 else 0
        
        cat_row_idx = len(table_data)
        cat_row_indices.append(cat_row_idx)
        table_styles_list.append(('BACKGROUND', (0, cat_row_idx), (-1, cat_row_idx), COLOR_CAT))
        table_styles_list.append(('TEXTCOLOR', (0, cat_row_idx), (-1, cat_row_idx), colors.white))
        table_styles_list.append(('FONTSIZE', (0, cat_row_idx), (-1, cat_row_idx), 9))
        table_styles_list.append(('SPAN', (0, cat_row_idx), (6, cat_row_idx)))
        
        cat_row = [
            Paragraph(f'<b>{cat["name"].upper()}</b>', ParagraphStyle('Cat', textColor=colors.white, fontSize=9)),
            '', '', '', '', '', '',
            Paragraph(f'<b>S/ {cat_total:,.2f}</b>', ParagraphStyle('CatT', textColor=colors.white, fontSize=9, alignment=TA_RIGHT)),
            Paragraph(f'<b>{pct_cat:.1f}%</b>', ParagraphStyle('CatP', textColor=colors.white, fontSize=9, alignment=TA_CENTER)),
        ]
        table_data.append(cat_row)
        
        for item in items:
            item_counter += 1
            pct_item = (item.get('total_cost', 0) / grand_total * 100) if grand_total > 0 else 0
            item_row = [
                item_counter,
                item.get('code', ''),
                item['description'],
                item.get('cost_type', '').capitalize(),
                item.get('unit', ''),
                f"{item.get('quantity', 0):,.3f}",
                f"S/ {item.get('unit_cost', 0):,.2f}",
                f"S/ {item.get('total_cost', 0):,.2f}",
                f"{pct_item:.1f}%",
            ]
            table_data.append(item_row)
    
    # Total general
    total_row_idx = len(table_data)
    table_styles_list.append(('BACKGROUND', (0, total_row_idx), (-1, total_row_idx), COLOR_YELLOW))
    table_styles_list.append(('FONTSIZE', (0, total_row_idx), (-1, total_row_idx), 10))
    table_styles_list.append(('SPAN', (0, total_row_idx), (6, total_row_idx)))
    
    table_data.append([
        Paragraph('<b>COSTO TOTAL GENERAL</b>', ParagraphStyle('GT', fontSize=10, alignment=TA_RIGHT, textColor=COLOR_DARK)),
        '', '', '', '', '', '',
        Paragraph(f'<b>S/ {grand_total:,.2f}</b>', ParagraphStyle('GTT', fontSize=10, alignment=TA_RIGHT, textColor=COLOR_DARK)),
        Paragraph('<b>100.0%</b>', ParagraphStyle('GTP', fontSize=10, alignment=TA_CENTER, textColor=COLOR_DARK)),
    ])
    
    t = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
    t.setStyle(TableStyle(table_styles_list))
    story.append(t)
    
    # Resumen financiero
    monthly_prod = structure.get('monthly_production') or 0
    sale_price = structure.get('sale_price') or 0
    
    if monthly_prod > 0:
        story.append(Spacer(1, 20))
        story.append(HRFlowable(width="100%", thickness=1, color=COLOR_BLUE))
        story.append(Paragraph("<b>RESUMEN FINANCIERO</b>",
            ParagraphStyle('RF', parent=styles['Heading2'], textColor=COLOR_DARK, fontSize=12)))
        
        costo_unit = grand_total / monthly_prod
        resumen_data = [
            ['Producción mensual', f"{monthly_prod:,.0f} unidades"],
            ['Costo total', f"S/ {grand_total:,.2f}"],
            ['Costo unitario', f"S/ {costo_unit:,.4f}"],
        ]
        if sale_price > 0:
            precio_sin_igv = sale_price / 1.18
            utilidad = precio_sin_igv - costo_unit
            margen = (utilidad / precio_sin_igv * 100) if precio_sin_igv > 0 else 0
            resumen_data += [
                ['Precio de venta (con IGV)', f"S/ {sale_price:,.2f}"],
                ['Precio sin IGV', f"S/ {precio_sin_igv:,.2f}"],
                ['Utilidad unitaria', f"S/ {utilidad:,.4f}"],
                ['Margen de utilidad', f"{margen:.1f}%"],
            ]
        
        t_res = Table(resumen_data, colWidths=[8*cm, 6*cm])
        t_res.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), COLOR_LIGHT),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ]))
        story.append(t_res)
    
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        f"<i>Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')} | Sistema de Costos Empresariales - Perú</i>",
        ParagraphStyle('Footer', fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    ))
    
    doc.build(story)
    buffer.seek(0)
    return buffer
